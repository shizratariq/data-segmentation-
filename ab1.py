from openpyxl import load_workbook
book=load_workbook('power.xlsx')
sheet=book['Trial1']

#initializing variables
k=0
res=[]
fsr1=[]
seg1=[]
end=0
#removing string value from array
for row in sheet.rows:
    str(row[0].value)
    fsr1.append(row[0].value)
fsr1.pop(0)
#removing none value from array
for val in fsr1:
    if val != None :
        res.append(val)

def segment(k):
    for i in range(k, len(res)):
        if res[i] > 100:
            seg1.append(res[i])
        elif len(seg1) > 40:
            end=i
            return end
            break

def write(j):
    import xlsxwriter

    workbook = xlsxwriter.Workbook('arrays.xlsx')
    worksheet = workbook.add_worksheet()
    row = 0
    for i in range(len(seg1)):
        worksheet.write(i, 0, seg1[i])
    workbook.close()

end= segment(k)
write(0)
